import pandas as pd
import numpy as np
import os
import openpyxl
import dateutil
import datetime
import carlao
from datetime import timedelta
import zipfile
from math import ceil, isnan
from pathlib import Path
from scipy import interpolate
import re
from shutil import rmtree
from sys import exit
import cargill
import shutil

#supressao avisos
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings('ignore')
#----------
#Versão V5.4
#####################################################################################################
#                           TEXT AND DATA REPLACEMENT / INVALID DATA CLEANING                       #
#####################################################################################################
dic_tmi = pd.DataFrame(np.array([
    ['ENGINE SPEED','RPM'],
    ['ENGINE POWER','POWER'],
    ['PERCENT LOAD','LOAD'],
    ['ENGINE LOAD','LOAD'],
    ['ENGINE TORQUE','TORQUE'],
    ['BRAKE MEAN EFF PRES','BMEP'],
    ['BRAKE SPEC FUEL CONSUMPTN','BSFC'],
    ['ISO BRAKE SPEC FUEL CONSUMPTN','ISOBSFC'],
    ['VOL FUEL CONSUMPTN','VFC'],
    ['ISO VOL FUEL CONSUMPTN','ISOVFC']
    ]
),columns=['de','para'])

events_replace_text = pd.DataFrame(np.array(
    [['Severity', 'Severity'], ['Type', 'Type'], ['Code', 'Code'], ['Source', 'Source'], ['Description', 'Description'],
     ['Sample Time', 'Timestamp'], ['Run Hours', 'Run_Hours'], ['Events', 'Event'], ['Disgnostics', 'Disgnostic']]
), columns=['de', 'para'])


data_replace_text = pd.DataFrame(np.array([
    ['Sample', 'Timestamp'], ['Engine Load', 'Load'], ['Engine Speed', 'RPM'],
    ['Engine Coolant', 'Coolant_Temp'], ['Oil Pressure', 'Oil_Press'], ['Oil Temperature', 'Oil_Temp'],
    ['Battery Voltage', 'Batt'],
    ['Boost Pressure', 'Boost'], ['Fuel Consumption Rate', 'Fuel_Rate'], ['Left Exhaust Temp', 'EXH_L'],
    ['Right Exhaust Temp', 'EXH_R'],
    ['Total Fuel', 'Total_Fuel'], ['Run Hours', 'SMH'], ['Fuel Pressure', 'Fuel_Press'],
    ['Crankcase Pressure', 'Crank_Press'],
    ['Latitude', 'Latitude'], ['Longitude', 'Longitude'], ['Vessel Speed', 'Vessel_Speed'],
    ['Diagnostic Status', 'DIAG_STATS'], ['Diagnostic Code - CID', 'CID'], ['Diagnostic SubCode - FMI', 'FMI'],
    ['Event Status', 'EVENT_STATS'], ['Event Code', 'EID']
]
), columns=['de', 'para'])

maintenance_list = ['Site', 'Asset', 'SH_by_Day', 'SMH_Calc', 'SMH', 'Fuel_by_Day', 'Total_Fuel_Calc', 'Total_Fuel',
                    'Next_Prev', 'Prev_Date', 'Next_OVH', 'OVH_Date', 'Per_by', 'Ovh_by']

std_param_list = ['Timestamp', 'Load', 'RPM', 'Coolant_Temp', 'Oil_Press', 'Oil_Temp', 'Batt', 'Boost', 'Fuel_Rate',
                  'EXH_L', 'EXH_R', 'Total_Fuel', 'Fuel_Press', 'Crank_Press', 'Total_Fuel_DIFF', 'SMH_DIFF',
                  'EXH_DIFF']

std_event_list = ['Severity', 'Type', 'Code', 'Source', 'Description', 'Timestamp', 'Run_Hours']

std_eventsum_list = ['Severity', 'Type', 'Code', 'Source', 'Description']

std_rpmhist_list = ['RPM Range (%)', 'Time (%)', 'Hours (h)', 'Asset', 'Site']

std_loadhist_list = ['Power Range (%)', 'Time (%)', 'Hours (h)', 'Asset', 'Site']

remove_prefix_list = ['dg1_', 'dg2_', 'dg3_', 'dg4_', 'ple_cat_', 'mca', 'mcp', 'bb', 'be', 'cn', '_']



#####################################################################################################
#                                           BASIC FUNCTIONS START                                   #
#####################################################################################################

def contacol(df, tx):
    coldf = df.columns.tolist()
    count = 0
    for col in coldf:
        count = count + len(re.findall(tx, col, flags=re.I))
    return count

def findpower(df,rpmval):
    x = df['RPM']
    y = df['POWER']
    f = interpolate.interp1d(x,y,fill_value='extrapolate')
    pwval = f(rpmval)
    return pwval

def powercalc(tmidf,engdf):
    df = replace_coltext_df(tmidf, dic_tmi)#.apply(pd.to_numeric, errors='ignore')

    for col in engdf.columns:
        if col != 'Timestamp':
          engdf[col] = pd.to_numeric(engdf[col], errors='ignore')

    for col in df.columns:
        if col != 'Timestamp':
          df[col] = pd.to_numeric(df[col], errors='ignore')

    if not 'LOAD' in tmidf.columns:
        
        if len(engdf['RPM']) > 1:
            engdf.drop(engdf[(engdf['RPM'] == 0.0) | (engdf['Load'] == 0.0) | 
                                (engdf['RPM'] < 0.0) | (engdf['Load'] < 0.0) | 
                                (engdf['RPM'] >2500) | (engdf['Load'] > 150)
                                ].index,inplace=True)
            engdf['MaxPower'] = findpower(df,engdf['RPM'])
            engdf['RealPower']=engdf['MaxPower']*engdf['Load']/100
            engdf['BSFC'] = engdf['Fuel_Rate']/engdf['RealPower']*f_density
            engdf['BSFC'] = engdf['BSFC'].apply(pd.to_numeric)
    else:
        if len(engdf['Load']) > 1:
            engdf.drop(engdf[(engdf['RPM'] == 0.0) | (engdf['Load'] == 0.0) | 
                                (engdf['RPM'] < 0.0) | (engdf['Load'] < 0.0) | 
                                (engdf['RPM'] >2500) | (engdf['Load'] > 150)
                                ].index,inplace=True)
            engdf['MaxPower'] = df['POWER'].max()
            engdf['RealPower'] = engdf['MaxPower']*engdf['Load']/100
            engdf['BSFC'] = engdf['Fuel_Rate']/engdf['RealPower']*f_density
            engdf['BSFC'] = engdf['BSFC'].apply(pd.to_numeric)
    return engdf

def openfilewb(input_file,ws):
    in_file = os.path.join(input_file)
    outwb = openpyxl.load_workbook(in_file)
    data = outwb[ws].values
    columns = next(data)[0:]
    df = pd.DataFrame(data,columns=columns)
    df.dropna(how='all',inplace=True)
    return df 

def cons_perfn(asset):
    perfseries = assetlistdf.loc[assetlistdf['Serial'] == asset]['Perf_Number'].tolist()
    return perfseries[0]

def limpadao(dir):
    for stuff in os.listdir(dir):
        coisa = os.path.join(dir, stuff)
        if os.path.isdir(coisa) == True:
            for stuff2 in os.listdir(coisa):
                if os.path.isdir(stuff2) == False:
                    if stuff2.endswith('output.csv'):
                        os.replace(os.path.join(coisa, stuff2), os.path.join(dir, stuff2))
    if keepfiles == 0:
        for file_name in os.listdir(dir):
            folder = os.path.join(dir, file_name)
            if os.path.isdir(folder) == True:
                os.rmdir(folder)

# Tenta criar o diretório. Caso já exista, passa adiante.
def checkdestiny(outputdirectory):
    try:
        os.mkdir(outputdirectory)
    except FileExistsError:
        pass

def removeprefix(text, plist):
    for ch in plist:
        if ch in text:
            text = text.replace(ch, "")
    return text

def replace_coltext_df(df, replacement_df):
    o = len(replacement_df.axes[0])
    r = len(replacement_df.axes[1])
    i = 0
    while i < o:
        df.columns = df.columns.str.replace(r'('+replacement_df.iloc[i, 0] + '.*$)', replacement_df.iloc[i, 1],regex=True)
        i += 1
    return df

def replace_text(df, replacement_df):
    o = len(replacement_df.axes[0])
    r = len(replacement_df.axes[1])
    i = 0
    while i < o:
        df.columns = df.columns.str.replace(r'(^.*' + replacement_df.iloc[i, 0] + '.*$)', replacement_df.iloc[i, 1],
                                            regex=True)
        i += 1
    return df

def csvfix(df, lista):
    dffix = df.copy(deep=True)
    for h in lista:
        if h not in dffix.columns:
            dffix[h] = ''
    return dffix

def csvfix2(colunas, linhas):
    dffix = pd.DataFrame(linhas, columns=colunas)
    return dffix

def dateparser(filein):
    date_parser = lambda x: dateutil.parser.parse(x, ignoretz=True)
    tabin = pd.read_csv(filein, dayfirst=False, parse_dates=['Sample Time'], date_parser=date_parser)
    tabin.to_csv(filein, encoding='utf-8-sig', index=True)

def getlistativos(file):
    # pega a aba 'Engine Event Summary' e cria um array com todos os números de série da coluna 'Unit Name', tirando a linha de totais.
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        global ws_Eng_Summ
        ws_Eng_Summ = wb['Engine Event Summary']
        data = ws_Eng_Summ.values
        df = pd.DataFrame(data, columns=next(data)[0:])
        df = df.loc[:, df.columns == 'Unit Name']
    return df['Unit Name']

def getlistasites(file):
    # pega a aba 'ASSET_LIST' e cria um array com todos os nomes de sítios.
    if file.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        global ws_Site_Summ
        ws_Site_Summ = wb['ASSET_LIST']
        data = ws_Site_Summ.values
        df = pd.DataFrame(data, columns=next(data)[0:])
        df = df.loc[:, df.columns == 'Vessel']
        df.drop_duplicates(subset=['Vessel'],inplace=True)
    return df['Vessel']

def concatenar(dir, name):
    outname = name + 'output.csv'
    for file_name in os.listdir(dir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(dir + '/' + outname, low_memory=False)
                b = pd.read_csv(dir + file_name, low_memory=False)
                a.columns = a.columns.str.replace(" ","_",regex=True)
                b.columns = b.columns.str.replace(" ","_",regex=True)
                df = pd.concat([a, b])
                df.drop_duplicates(subset=['Timestamp', 'Asset'], inplace=True, keep='last')
                # df.dropna(how='all', axis=1, inplace=True)
                df.to_csv(dir + '/' + outname, encoding='utf-8-sig', index=False)
    if keepfiles == 0:
        for file_name in os.listdir(dir):
            if file_name != outname:
                os.remove(dir + file_name)
    return df

def concatenarev(dir, name):
    outname = name + 'output.csv'
    for file_name in os.listdir(dir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(dir + '/' + outname, low_memory=False)
                b = pd.read_csv(dir + file_name, low_memory=False)
                a.columns = a.columns.str.replace(" ","_",regex=True)
                b.columns = b.columns.str.replace(" ","_",regex=True)
                df = pd.concat([a, b])
                df.drop_duplicates(inplace=True, keep='last')
                # df.dropna(how='all', axis=1, inplace=True)
                df.to_csv(dir + '/' + outname, encoding='utf-8-sig', index=False)
                
    if keepfiles == 0:
        for file_name in os.listdir(dir):
            if file_name != outname:
                os.remove(dir + file_name)

    return df

def concatenar_profile(hdir, columns, name):
    outname = name + 'output.csv'
    try:
        
        os.remove(hdir + '/' + outname)
        cdf1 = pd.DataFrame(columns=columns)
        cdf1.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf1 = pd.DataFrame(columns=columns)
        cdf1.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)

    for file_name in os.listdir(hdir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(hdir + '/' + outname, low_memory=False)
                b = pd.read_csv(hdir + file_name, low_memory=False)
                dfs = pd.concat([a, b])
                dfs.drop_duplicates(subset=None, inplace=True, keep='last')
                # dfs.dropna(how='all', axis=1, inplace=True)
                dfs.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
    if keepfiles == 0:
        for file_name in os.listdir(hdir):
            if file_name != outname:
                os.remove(hdir + file_name)
    
    return dfs

def concatenar_study(hdir, name):
    outname = name + 'output.csv'
    try:
        os.remove(hdir + '/' + outname)
        cdf = pd.DataFrame(columns=['Data'])
        cdf.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=['Data'])
        cdf.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)

    for file_name in os.listdir(hdir):
        if file_name != outname:
            if file_name != scriptname:
                a = pd.read_csv(hdir + '/' + outname, low_memory=False)
                b = pd.read_csv(hdir + file_name, low_memory=False)
                df = pd.concat([a, b])
                df.drop_duplicates(subset=None, inplace=True, keep='last')
                # df.dropna(how='all', axis=1, inplace=True)
                df.to_csv(hdir + '/' + outname, encoding='utf-8-sig', index=False)
    if keepfiles == 0:
        for file_name in os.listdir(hdir):
            if file_name != outname:
                os.remove(hdir + file_name)
    return df

def findsitename(asset):
    wb_info = openpyxl.load_workbook(os.path.join(infodir, 'ASSET_INFO.xlsx'))
    ws_a_list = wb_info['ASSET_LIST']
    data = ws_a_list.values
    cuslist = pd.DataFrame(data, columns=next(data)[0:])
    site_name = cuslist.loc[cuslist['Serial'] == asset]['Vessel'].tolist()
    return site_name[0]

def findsiteassets(site):
    wb_info = openpyxl.load_workbook(os.path.join(infodir, 'ASSET_INFO.xlsx'))
    ws_a_list = wb_info['ASSET_LIST']
    data = ws_a_list.values
    df = pd.DataFrame(data, columns=next(data)[0:])
    slist = df.loc[df['Vessel'] == site]['Serial'].tolist()
    return slist

def findmodel(asset):
    df = pd.read_excel(os.path.join(infodir, 'ASSET_INFO.xlsx'), sheet_name='ASSET_LIST')
    model = df.loc[df['Serial'] == asset]['Model'].tolist()
    return model[0]

def renamecol(df, sn):
    dflistparm = dflistparmraw.loc[dflistparmraw['SN']==sn]
    dflistparm.reset_index(drop=True, inplace=True)
    i=0
    while i < len(dflistparm):
        col1 = dflistparm.loc[i, 'Nome da coluna']
        if col1 in df.columns:
            x = dflistparm.loc[i, 'Renomear para']
            df = df.rename(columns={col1: x})
            i += 1
    return df

def delcol(df):
    
    columns = dflistcoldel.values.tolist()
    columns = list(np.concatenate(columns).flat)
    
    for col in columns:
        if df.columns.tolist().count(col)>0:
            df = df.drop(col, axis = 1)
    return df

def delalerts(df):
    
    listalerts = dflistalertdel.values.tolist()
    listalerts = list(np.concatenate(listalerts).flat)
    for i in listalerts:
        df = df[df['Code'] != i]
        
    return df

        
#####################################################################################################
#                                           MANUTENÇÃO                                              #
#####################################################################################################
# ele vai ter que receber o nome do ativo para filtrar no dataframe esse ativo 
# e aí fazer os cálculos de manutenção!!!!

def maintenanceoutput(dfoutput, lastused, asset_sn, datasetvazio):
    site_name = findsitename(asset_sn)

    fcd, fct, totfc = dictfuelday[asset_sn]
    shd, sht, totsh = dictsmhday[asset_sn]

    print('Ativo:', str(asset_sn))
    print('Consumo por dia:', str(fcd), 'Consumo Total do período:', str(fct))
    print('Horas de serviço por dia:', str(shd), 'Horas de serviço do período:', str(sht))

    if isnan(fcd) and isnan(shd):
        next_preventiva = np.nan
        next_preventiva_day = np.nan
        next_overhaul = np.nan
        next_overhaul_day = np.nan
        manper_by = np.nan
        manovh_by = np.nan
    else:
        try:
            (next_preventiva, next_preventiva_day,
            next_overhaul,next_overhaul_day,
                manper_by, manovh_by)  = maintcalc(totsh, shd, totfc, fcd, lastused, asset_sn)
        except Exception as erro:
            print('Erro na Manutenção!!!')
            print(erro)
            next_preventiva = np.nan
            next_preventiva_day = np.nan
            next_overhaul = np.nan
            next_overhaul_day = np.nan
            manper_by = np.nan
            manovh_by = np.nan


    print('Proxima manutenção preventiva:', str(next_preventiva), 'horas.', 'Data:', str(next_preventiva_day), '  Método: ', manper_by)
    print('Proximo overhaul:', str(next_overhaul), '.', 'Data:', str(next_overhaul_day), '  Método: ', manovh_by)
    print(' ')

    df = pd.DataFrame([[site_name, asset_sn, shd, sht, totsh, fcd, fct, totfc, next_preventiva, next_preventiva_day,
                        next_overhaul, next_overhaul_day, manper_by, manovh_by]],
                      columns=maintenance_list)
    return df

def fuelcalc(df, assetname):
    global fuelbyday
    global fuelcons
    global totalfuel
    global ndays

    df = df.query("Asset == @assetname")
    #df = df.query("RPM > 0")
    if not df.shape[0] < 1:
        if 'Total_Fuel' in df.columns:
            try:
                ndays = np.timedelta64(pd.to_datetime(df['Timestamp']).max()-pd.to_datetime(df['Timestamp']).min(),'h').astype(int)/24
            except:
                ndays = 1

            totalfuel = round(pd.to_numeric(df['Total_Fuel'].max()), 0)
            try:
                fuelcons = pd.to_numeric(df['Total_Fuel'].max() - df['Total_Fuel'].min())
                fuelcons = round(fuelcons, 0)
            except:
                fuelcons = np.nan

            if np.isnan(fuelcons) or ndays == 1 or np.isnan(ndays) or ndays == 0 or fuelcons == 0:
                fuelcons = np.nan
                fuelbyday = np.nan
            else:
                try:
                    fuelbyday = int(int(fuelcons) / ndays)
                    fuelbyday = round(fuelbyday, 0)
                except ValueError:
                    fuelbyday = np.nan
        else:
            fuelcons = np.nan
            fuelbyday = np.nan
            totalfuel = np.nan
    else:
        fuelcons = np.nan
        fuelbyday = np.nan
        totalfuel = np.nan
    return [fuelbyday, fuelcons, totalfuel]

def smhcalc(df, assetname):
    global smhbyday
    global servicehr
    global totalsh
    global ndays

    df = df.query("Asset == @assetname")
    #df = df.query("RPM > 0")
    if not df.shape[0] < 1:
        if 'SMH' in df.columns:
            try:
                ndays = np.timedelta64(pd.to_datetime(df['Timestamp']).max() - pd.to_datetime(df['Timestamp']).min(),
                                   'h').astype(int) / 24
            except:
                ndays = 1

            totalsh = round(pd.to_numeric(df['SMH'].max()), 0)
            try:
                servicehr = pd.to_numeric(df['SMH'].max() - df['SMH'].min())
                servicehr = round(servicehr, 0)
            except:
                servicehr = np.nan

            if np.isnan(servicehr) or ndays == 1 or np.isnan(ndays) or ndays == 0 or servicehr == 0:
                servicehr = np.nan
                smhbyday = np.nan
            else:
                try:
                    smhbyday = ceil(servicehr / ndays)
                    smhbyday = round(smhbyday, 0)
                except ValueError:
                    smhbyday = np.nan
        else:
            servicehr = np.nan
            smhbyday = np.nan
            totalsh = np.nan
            
    else:
        servicehr = np.nan
        smhbyday = np.nan
        totalsh = np.nan
    return [smhbyday, servicehr, totalsh]

# next_preventiva, next_preventiva_day = manutcalc(sht, shd, lastused, asset_sn, datasetvazio)
def maintcalc(lastsmh, hday, lastfuel, fday, lastdayused, sn):

    try:
        dfmetodo = dfmainte.loc[dfmainte['SN'] == sn]
        dfmetodo.reset_index(drop=True, inplace=True)
        metodo = dfmetodo.loc[0,'Médodo']

        if dfmetodo.shape[0] < 1:
            if metodo == 'SMH':
                fday = np.nan
            elif metodo == 'Fuel':
                hday = np.nan

    
    except:
        pass

    # PEGA O PREFIXO DE SÉRIE PARA BUSCAR NA PLANILHA ONDE ESTÃO OS INTERVALOS DE MANUTENÇÃO
    try:
        manplan = pd.read_excel(os.path.join(infodir + '/MAINTENANCE_PLAN.xlsx'), sheet_name ='By Model')
        # manplan = manplan.apply(pd.to_numeric, errors='coerce')
        manplan.loc[manplan['Maintenance Name'].notnull(),'Maintenance Name'] = manplan['Maintenance Name'].astype(str)
        manplan = manplan.dropna(how='all')
        model = findmodel(sn)
        manplan = manplan.loc[manplan['Model'] == model]
    except FileNotFoundError:
        nextpername = np.nan
        nextperday = np.nan
        nextovhname = np.nan
        nextovhday = np.nan
        manper_by = np.nan
        manovh_by = np.nan
        return nextpername, nextperday, nextovhname, nextovhday, manper_by, manovh_by

    # print('***interno função manutenção***')    
    manplan = manplan.apply(pd.to_numeric, errors='ignore')
    lastsmh = float(lastsmh)
    hday = float(hday)
    lastfuel = float(lastfuel)
    fday = float(fday)
    
    lastdayused = pd.to_datetime(lastdayused)


    
    try:
        manshift = pd.read_excel(os.path.join(infodir + '/MAINTENANCE_SHIFT.xlsx'))
        manshift.loc[manshift['Maintenance Name'].notnull(),'Maintenance Name'] = manshift['Maintenance Name'].astype(str)
        # manutenção realizada
        manshift = manshift.loc[manshift['SN'] == sn]
        # print(forcedman)
        if manshift.shape[0] < 1:
            #mandate = np.nan
            manfuel = 0
            mansmh = 0
            manter = 'Não'
            manshift_name = np.nan
        else:
            # dia da realização
            manshift.reset_index(drop=True, inplace=True)
            manfuel = float(manshift.loc[0, 'Total Fuel (L)'])
            mansmh = float(manshift.loc[0, 'Run Hours'])
            manter = manshift.loc[0,'Manter']
            manshift_name = str(manshift.loc[0,'Maintenance Name'])
            mandate = manshift.loc[0,'Date']
           
    except KeyError:
        #mandate = np.nan
        manfuel = 0
        mansmh = 0
    except FileNotFoundError:
        #mandate = np.nan
        manfuel = 0
        mansmh = 0
    
    
    manper = manplan.loc[manplan['Maintenance Type'] == 'Periódica']
    manovh = manplan.loc[manplan['Maintenance Type'] == 'Overhaul']
    
    ultovhsmh = manovh['Target SMH'].max()
    ultovhfuel = manovh['Target Fuel (L)'].max()

    
    if isinstance(manshift_name, str) or isinstance(manshift_name, str):
        stdmainsmh = manplan.loc[manplan['Maintenance Name']==manshift_name,'Target SMH'].tolist()[0]
        stdmainfuel = manplan.loc[manplan['Maintenance Name']==manshift_name,'Target Fuel (L)'].tolist()[0]
    
        if isnan(lastsmh):
            lastsmh = 0
        if isnan(lastfuel):
            lastfuel = 0
        if isnan(fday):
            fday = 0
        if isnan(hday):
            hday = 0
    
        if isnan(mansmh):
            daydiff = (lastdayused - mandate.days)*-1
            mansmh = lastsmh - hday*daydiff
            
        if isnan(manfuel):
            daydiff = (lastdayused - mandate).days
            manfuel = lastfuel - fday*daydiff
        
        if stdmainsmh != ultovhsmh:
            calsmh = lastsmh - (mansmh - ultovhsmh*int(mansmh/ultovhsmh) - stdmainsmh)
            calfuel = lastfuel - (manfuel - ultovhfuel*int(manfuel/ultovhfuel) - stdmainfuel)
        else:
            calsmh = lastsmh - (mansmh - ultovhsmh*int((mansmh/ultovhsmh)-1) - stdmainsmh)
            calfuel = lastfuel - (manfuel - ultovhfuel*int((manfuel/ultovhfuel)-1) - stdmainfuel)
    else:
        calsmh = lastsmh
        calfuel = lastfuel
    
    if isnan(calsmh):
        calsmh = 0
    if isnan(calfuel):
        calfuel = 0
    
    #Cálculo das periódicas - SMH
    ultper = manper['Target SMH'].max()
    ncyclesmh = calsmh/ultovhsmh
    calsmhper = calsmh - ultovhsmh*int(ncyclesmh)
    manper = manper.sort_values(by=['Target SMH'])
    if calsmhper> ultper:
        calsmhper = calsmh - ultovhsmh*(int(ncyclesmh)+1)
    for per in manper['Target SMH']:
        if per >= calsmhper:
            nextpersmh = manper.loc[manper['Target SMH'] == per]
            break
    
    #Cálculo das periódicas - Fuel
    
    ultper = manper['Target Fuel (L)'].max()
    ncyclefuel = calfuel/ultovhfuel
    calfuelper = calfuel - ultovhfuel*int(ncyclefuel)
    manper = manper.sort_values(by=['Target Fuel (L)'])
    if calfuelper> ultper:
        calfuelper = calfuel - ultovhfuel*(int(ncyclefuel)+1)
    for per in manper['Target Fuel (L)']:
        if per >= calfuelper:
            nextperfuel = manper.loc[manper['Target Fuel (L)'] == per]
            break
    
    #Cálculo das periódicas - Final
    try:
        persmh = nextpersmh.loc[:,'Target SMH']
    except ValueError:
        persmh = pd.DataFrame.from_dict({'Model':np.NaN,'Maintenance Name':np.NaN,'Maintenance Type':np.NaN, 'Target Fuel (L)':np.NaN,'Target SMH':np.NaN})
    try:
        perfuel = nextperfuel.loc[:,'Target Fuel (L)']
    except ValueError:
        perfuel = pd.DataFrame.from_dict({'Model':np.NaN,'Maintenance Name':np.NaN,'Maintenance Type':np.NaN, 'Target Fuel (L)':np.NaN,'Target SMH':np.NaN})
        
    if (hday == 0 or isnan(hday) == True) and (fday == 0 or isnan(fday) == True):
        nextpername = np.nan
        nextperday = np.nan
        nextovhname = np.nan
        nextovhday = np.nan
        manper_by = np.nan 
        manovh_by = np.nan
        return nextpername, nextperday, nextovhname, nextovhday, manper_by, manovh_by
    
    if hday == 0 or isnan(hday) == True:
        ncyclesmh = ncyclefuel - 1
    elif fday == 0 or isnan(fday) == True:
        ncyclefuel = ncyclesmh - 1
        
    if ncyclesmh >= ncyclefuel:
        manper_by = 'SMH'
        smh_per_remain = persmh - calsmhper
        ndayper = ceil(smh_per_remain/hday)
        nextpername = nextpersmh.loc[:,'Maintenance Name']
        
    else:
        manper_by = 'Fuel'
        fuel_per_remain = perfuel - calfuelper
        ndayper = ceil(fuel_per_remain/fday)
        nextpername = nextperfuel.loc[:,'Maintenance Name']
    
    try: 
        nextperday = lastdayused + pd.DateOffset(days=ndayper)
    except:
        nextperday = np.nan
    
    #Cálculo dos Overhaul - SMH
    
    ultovh = manovh['Target SMH'].max()
    ncyclesmh = calsmh/ultovh
    calsmhovh = calsmh - ultovh*int(ncyclesmh)
    manovh = manovh.sort_values(by=['Target SMH'])
    for ovh in manovh['Target SMH']:
        if ovh > calsmhovh:
            nextovhsmh = manovh.loc[manovh['Target SMH'] == ovh]
            break
    
    #Cálculo dos Overhaul - Fuel
    
    ultovh = manovh['Target Fuel (L)'].max()
    ncyclefuel = calfuel/ultovh
    calfuelovh = calfuel - ultovh*int(ncyclefuel)
    manovh = manovh.sort_values(by=['Target Fuel (L)'])
    for ovh in manovh['Target Fuel (L)']:
        if ovh > calfuelovh:
            nextovhfuel = manovh.loc[manovh['Target Fuel (L)'] == ovh]
            break
    
    #Cálculo dos Overhaul - Final
    ovhsmh = nextovhsmh.loc[:,'Target SMH']
    ovhfuel = nextovhfuel.loc[:,'Target Fuel (L)']
    
    try:
        ovhsmh = nextovhsmh.loc[:,'Target SMH']
    except ValueError:
        ovhsmh = np.NaN
    try:
        ovhfuel = nextovhfuel.loc[:,'Target Fuel (L)']
    except ValueError:
        ovhfuel = np.NaN

    
    if hday == 0 or isnan(hday) == True:
        ncyclesmh = ncyclefuel - 1
    elif fday == 0 or isnan(fday) == True:
        ncyclefuel = ncyclesmh - 1
     
    if ncyclesmh >= ncyclefuel:
        manovh_by = 'SMH'
        smh_ovh_remain = ovhsmh - calsmhovh
        ndayovh = ceil(smh_ovh_remain/hday)
        nextovhname = nextovhsmh.loc[:,'Maintenance Name']
        
    else:
        manovh_by = 'Fuel'
        fuel_ovh_remain = ovhfuel - calfuelovh
        ndayovh = ceil(fuel_ovh_remain/fday)
        nextovhname = nextovhfuel.loc[:,'Maintenance Name']
    
    try:
        nextovhday = lastdayused + pd.DateOffset(days=ndayovh)
    except:
        nextovhday = np.nan
    
    if manter == 'Manter':
        #contador i não está funcionando
        mainname = manshift.loc[0,'Maintenance Name']
        manovh.reset_index(drop=True, inplace=True)
        i=1
        for name in manovh['Maintenance Name']:
            if mainname == name:
                break
            else:
                i=i+1
        try:
            nextovh = manovh.loc[i,:]
        except KeyError:
            nextovh = manovh.loc[0,:]
        
        ovhsmh = float(nextovh['Target SMH'])
        ovhfuel = float(nextovh['Target Fuel (L)'])
        smh_ovh_remain = ovhsmh - calsmhovh
        fuel_ovh_remain = ovhfuel - calfuelovh
        
        if fday != 0:
            ndayovhfuel = ceil(fuel_ovh_remain/fday)
        else:
            ndayovhfuel = np.nan
        
        if hday != 0:
            ndayovhsmh = ceil(smh_ovh_remain/hday)
        else:
            ndayovhsmh = np.nan
        
        nextovhname = nextovh['Maintenance Name']
        
        try:
        
            if ndayovhsmh <= ndayovhfuel:
                nextovhday = lastdayused + pd.DateOffset(days=ndayovhsmh)
            else:
                nextovhday = lastdayused + pd.DateOffset(days=ndayovhfuel)
        
        except:
            
            if ndayovhfuel == np.nan:
                nextovhday = lastdayused + pd.DateOffset(days=ndayovhsmh)
            elif ndayovhsmh == np.nan:
                nextovhday = lastdayused + pd.DateOffset(days=ndayovhfuel)
            
        return nextpername.tolist()[0], nextperday, nextovhname, nextovhday, manper_by, manovh_by
            
                
                    
            
    
    return nextpername.tolist()[0], nextperday, nextovhname.tolist()[0], nextovhday, manper_by, manovh_by
#####################################################################################################
#                                           HISTOGRAMAS                                             #
#####################################################################################################

def load_histogram(dataframe, rawdf, a_sn):
    perc = []
    PMax = 110
    Pdiv = 10
    h = (PMax // Pdiv) + 1
    n = 0
    for b in range(h):
        perc.append(n)
        n += 10
    dataframe["Load"] = dataframe["Load"].apply(pd.to_numeric)
    dataframe["RPM"] = dataframe["RPM"].apply(pd.to_numeric)
    engine_pw = dataframe.query("RPM > 0")["Load"]
    serieshist1 = engine_pw.value_counts(sort=False, bins=perc, normalize=True).round(4) * 100
    dfhist1 = serieshist1.rename_axis('unique_values').reset_index(name='counts')
    dfhist1 = engine_pw.value_counts(sort=False, bins=perc, normalize=True).reset_index().rename(
        columns={'index': 'bin'})
    dfhist1['Power Range (%)'] = ['({}, {})'.format(x.left, x.right) for x in dfhist1['bin']]
    dfhist1.columns = dfhist1.columns.str.replace('Load', 'Time (%)')
    dfhist1['Time (%)'] = dfhist1['Time (%)'].round(4) * 100
    dfhist1.drop(['bin'], axis=1, inplace=True)
    dfhist1['Power Range (%)'] = dfhist1['Power Range (%)'].map(
        {'(-0.001, 10.0)': '0-10%', '(10.0, 20.0)': '10-20%', '(20.0, 30.0)': '20-30%', '(30.0, 40.0)': '30-40%',
         '(40.0, 50.0)': '40-50%', '(50.0, 60.0)': '50-60%', '(60.0, 70.0)': '60-70%', '(70.0, 80.0)': '70-80%',
         '(80.0, 90.0)': '80-90%', '(90.0, 100.0)': '90-100%', '(100.0, 110.0)': '100-110%'}, na_action=None)

    if 'SMH' in rawdf.columns:
        rawdf["SMH"] = rawdf["SMH"].apply(pd.to_numeric)
        lasth = rawdf['SMH'].max()
        firsth = rawdf['SMH'].min()
        periodh = lasth - firsth
        dfhist1['Hours (h)'] = dfhist1['Time (%)'] * periodh / 100
        dfhist1 = dfhist1[['Power Range (%)', 'Time (%)', 'Hours (h)']]
    else:
        dfhist1 = dfhist1[['Power Range (%)', 'Time (%)']]
    outdir = os.path.join(destinationfolder, loadhistdir)
    checkdestiny(outdir)
    histout = outdir + a_sn + '_LOADPROFILE' + '.csv'

    listaaa = ['Power Range (%)', 'Time (%)', 'Hours (h)']
    rowws = [['0-10%', 0, 0], ['10-20%', 0, 0], ['20-30%', 0, 0], ['30-40%', 0, 0], ['50-60%', 0, 0], ['70-80%', 0, 0],
             ['90-100%', 0, 0], ['100-110%', 0, 0]]
    ndf = csvfix(dfhist1, listaaa)

    if ndf.shape[0] < 1:
        df = csvfix2(listaaa, rowws)
        df['Asset'] = asset_sn
        df['Site'] = findsitename(asset_sn)
        df.reset_index(drop=True, inplace=True)
        # df = df.transpose()
        df.to_csv(histout, encoding='utf-8-sig', index=False)
    else:
        ndf['Asset'] = asset_sn
        ndf['Site'] = findsitename(asset_sn)
        ndf.reset_index(drop=True, inplace=True)
        # ndf = ndf.transpose()
        ndf.to_csv(histout, encoding='utf-8-sig', index=False)

def rpm_histogram(dataframe, rawdf, a_sn):
    rawdf['RPM'] = rawdf['RPM'].apply(pd.to_numeric)
    # HISTOGRAMAS DE ROTAÇÃO
    perc = []
    PMax = 1900
    Pdiv = 100
    h = 20
    n = 0
    for b in range(h):
        perc.append(n)
        n += 100
    global dfhist2
    dataframe['RPM'] = dataframe['RPM'].apply(pd.to_numeric)
    engine_rpm = dataframe.query("RPM > 0")['RPM']
    serieshist2 = engine_rpm.value_counts(sort=False, bins=perc, normalize=True).round(4) * 100
    dfhist2 = serieshist2.rename_axis('unique_values').reset_index(name='counts')
    dfhist2 = engine_rpm.value_counts(sort=False, bins=perc, normalize=True).reset_index().rename(
        columns={'index': 'bin'})
    dfhist2['Range'] = ['({}, {})'.format(x.left, x.right) for x in dfhist2['bin']]
    dfhist2.columns = dfhist2.columns.str.replace('RPM', 'Time (%)')
    dfhist2.columns = dfhist2.columns.str.replace('Range', 'RPM Range (%)')
    dfhist2['Time (%)'] = dfhist2['Time (%)'].round(4) * 100
    dfhist2.drop(['bin'], axis=1, inplace=True)
    dfhist2['RPM Range (%)'] = dfhist2['RPM Range (%)'].map(
        {'(-0.001, 100.0)': '0-100', '(100.0, 200.0)': '100-200', '(200.0, 300.0)': '200-300',
         '(300.0, 400.0)': '300-400', '(400.0, 500.0)': '400-500', '(500.0, 600.0)': '500-600',
         '(600.0, 700.0)': '600-700', '(700.0, 800.0)': '700-800', '(800.0, 900.0)': '800-900',
         '(900.0, 1000.0)': '900-1000', '(1000.0, 1100.0)': '1000-1100', '(1100.0, 1200.0)': '1100-1200',
         '(1200.0, 1300.0)': '1200-1300', '(1300.0, 1400.0)': '1300-1400', '(1400.0, 1500.0)': '1400-1500',
         '(1500.0, 1600.0)': '1500-1600', '(1600.0, 1700.0)': '1600-1700', '(1700.0, 1800.0)': '1700-1800',
         '(1800.0, 1900.0)': '1800-1900'}, na_action=None)
    if 'SMH' in rawdf.columns:
        lasth = rawdf['SMH'].max()
        firsth = rawdf['SMH'].min()
        periodh = lasth - firsth
        dfhist2['Hours (h)'] = dfhist2['Time (%)'] * periodh / 100
        dfhist2 = dfhist2[['RPM Range (%)', 'Time (%)', 'Hours (h)']]
    else:
        dfhist2 = dfhist2[['RPM Range (%)', 'Time (%)']]

    outdir = os.path.join(destinationfolder, rpmhistdir)
    checkdestiny(outdir)
    histout = outdir + a_sn + '_RPMPROFILE' + '.csv'

    listbbb = ['RPM Range (%)', 'Time (%)', 'Hours (h)']

    df = csvfix(dfhist2, listbbb)
    df['Asset'] = asset_sn
    df['Site'] = findsitename(asset_sn)
    df.reset_index(drop=True, inplace=True)
    # df = df.transpose()
    df.to_csv(histout, encoding='utf-8-sig', index=False)

#####################################################################################################
#                            ESTUDOS DE UTILIZAÇÃO E TAXAS DE CONSUMO                               #
#####################################################################################################

def genloadxhour(df, base_path, loadtable, fratetable):
    vnames = df.columns
    for vname in vnames:
        dfvname = [x for x in df[vname] if str(x) != 'None']
        for a in range(len(dfvname)):
            if a == 0:
                ltv = pd.concat([loadtable.loc[:, loadtable.columns.isin(['Timestamp'])],
                                 loadtable.loc[:, loadtable.columns.str.contains(dfvname[a], case=False)]], axis=1)
                a += 1
            else:
                ltv = pd.concat([ltv, loadtable.loc[:, loadtable.columns.str.contains(dfvname[a], case=False)]], axis=1)

        if len(ltv.columns) != 0:
            ltv = ltv.replace(0, np.nan)
            ltv['QNT_SIM'] = ltv.apply(lambda x: x.notnull().sum(), axis='columns')

            fin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(dfvname) + 1)], ignore_index=True)
            l = []
            for i in range(len(dfvname) + 1):
                l.append(ltv[ltv['QNT_SIM'] == (i)].count()['QNT_SIM'] / len(ltv['QNT_SIM']) * 100)
            fin['%'] = l
            finfile = os.path.join(base_path) + '/' + vname + '_LOAD_RESUME' + '.csv'
            fin.to_csv(finfile, encoding='utf-8-sig', index=False)

            ltvd = ltv.resample('1D').mean().round(0)
            for x in range(len(ltv.columns) - 1):
                ltvd['UT' + str(x + 1) + 'E'] = (
                        (ltv['QNT_SIM'] == (x + 1)).resample('1D').sum().astype(int) / 24 * 100).round(1)

            ltvd.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvdfile = os.path.join(base_path) + '/' + vname + '_LOAD_STUDY' + '.csv'
            ltvd.to_csv(ltvdfile, encoding='utf-8-sig', index=True)

        # transforma v1 em tabela de taxa de consumo / hora
        for a in range(len(dfvname)):
            if a == 0:
                rtv = pd.concat([fratetable.loc[:, fratetable.columns.isin(['Timestamp'])],
                                 fratetable.loc[:, fratetable.columns.str.contains(dfvname[a], case=False)]], axis=1)
                a += 1
            else:
                rtv = pd.concat([rtv, fratetable.loc[:, fratetable.columns.str.contains(dfvname[a], case=False)]],
                                axis=1)

        rtvb = rtv.copy()
        rtvb = rtvb.replace(0, np.nan)
        rtvb['AVG'] = rtvb.sum(numeric_only=True, axis=1)

        if len(rtv.columns) != 0:
            rtvfile = os.path.join(base_path) + '/' + vname + '_FR_STUDY' + '.csv'
            rtv = rtv.replace(0, np.nan)
            rtvD = rtv.resample('1D').mean().round(0)
            rtvD.to_csv(rtvfile, encoding='utf-8-sig', index=True)

            rtv['QNT_SIM'] = rtv.apply(lambda x: x.notnull().sum(), axis='columns')

            rin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(dfvname) + 1)], ignore_index=True)
            l = []
            for i in range(len(dfvname) + 1):
                l.append(rtvb[rtv['QNT_SIM'] == (i)].mean()['AVG'])
            rin['L/hr'] = l
            rinfile = os.path.join(base_path) + '/' + vname + '_FR_RESUME' + '.csv'
            rin.to_csv(rinfile, encoding='utf-8-sig', index=False)

#####################################################################################################
#                                           ROTINAS                                                 #
#####################################################################################################

def rotinas(dataframe):
    global dfclean
    #print('Agora ', asset_sn)
    dataframe['Asset'] = asset_sn
    dataframe['Site'] = findsitename(asset_sn)
    perfn = cons_perfn(asset_sn)
    print('Abrindo ativo:', str(asset_sn), '\nPerformance Number: ' + perfn)
    
    try:
        for col in dataframe.columns:
            if col != 'Timestamp':
                #print(col)
                dataframe[col] = pd.to_numeric(dataframe[col], errors='ignore')
    
    except:
        print('\nErro na coluna:',col, 'no ativo',asset_sn)
        print('Defina a coluna certa no ConfigScript para continuar')
        exit()
    

    dfclean = dataframe.replace(list_invalid_data, '')
    
    if not dfclean.shape[0] < 1:
        if 'SMH' in dfclean.columns:
            dfclean["SMH"] = dfclean["SMH"].apply(pd.to_numeric)
            dfclean.loc[dfclean['SMH'] == 0,'SMH'] = np.nan
            dfclean['SMH'].interpolate(inplace=True)
            dfclean['Date'] = dfclean['Timestamp'].dt.date
            dfbyday = dfclean.groupby(dfclean['Date']).agg(SMH_DIFF=("SMH", lambda x: max(x) - min(x)))
            dfclean = dfclean.merge(dfbyday, how='left',on='Date')
            dfclean = dfclean.drop('Date', axis =1)
        else:
            dfclean['SMH_DIFF'] = np.nan
            dfclean['SMH'] = np.nan
            print(asset_sn,'Sem dados de SMH')
            
        if 'Total_Fuel' in dfclean.columns:
            dfclean = dfclean.drop(dfclean[(dfclean['Total_Fuel'] == 0)].index)
            dfclean["Total_Fuel"] = dfclean["Total_Fuel"].apply(pd.to_numeric)
            dfclean.loc[dfclean['Total_Fuel'] == 1,'Total_Fuel'] = np.nan
            dfclean['Total_Fuel'].interpolate(inplace=True)
            dfclean['Date'] = dfclean['Timestamp'].dt.date
            dfbyday = dfclean.groupby(dfclean['Date']).agg(Total_Fuel_DIFF=("Total_Fuel", lambda x: max(x) - min(x)))
            dfclean = dfclean.merge(dfbyday, how='left',on='Date')
            dfclean = dfclean.drop('Date', axis =1)
        else:
            dfclean['Total_Fuel_DIFF'] = np.nan
            print(asset_sn,'Sem dados de Total Fuel')

        if 'RPM' in dfclean.columns:
            MeanUR = dfclean['Timestamp'].diff().mean()
            Mintime = datetime.timedelta(minutes=10)
            if MeanUR < Mintime:
                dfclean['RPM'].interpolate(limit=1, inplace=True)
            dfclean = dfclean.dropna(axis=0, subset=['RPM'])
            dfclean = dfclean.drop(dfclean[(dfclean['RPM'] == 0)].index)
            rpm_histogram(dfclean, dataframe, asset_sn)
        else:
            print(asset_sn,'Sem dados de RPM')

        if 'Load' in dfclean.columns:
            dfclean = dfclean.dropna(axis=0, subset=['Load'])
            try:
                dfclean = dfclean.drop(dfclean[(dfclean['RPM'] == 0) & (dfclean['Load'] == 0)].index)
            except KeyError:
                dfclean = dfclean.drop(dfclean[(dfclean['Load'] == 0)].index)
            try:    
                tmidf = openfilewb(os.path.join(infodir, 'TMI_INFO.xlsx'),perfn)
                dfclean = powercalc(tmidf,dfclean)
            except KeyError:
                print('Sem dados de performance. Alguns calculos nao serao realizados, incluindo BSFC e Potencia Real.')
            load_histogram(dfclean, dataframe, asset_sn)
        else:
            print(asset_sn,'Sem dados de Load Factor')


        if all(pd.Series(['EXH_L', 'EXH_R']).isin(dfclean.columns)):
            dfclean["EXH_L"] = dfclean["EXH_L"].apply(pd.to_numeric)
            dfclean["EXH_R"] = dfclean["EXH_R"].apply(pd.to_numeric)
            dfclean['EXH_DIFF'] = abs(dfclean.EXH_L - dfclean.EXH_R)
        else:
            dfclean['EXH_DIFF'] = np.nan

        dfclean.set_index(pd.DatetimeIndex(dfclean['Timestamp']), inplace=True)
        print('Analise Ok.\n')
    else:
        print('Datalog vazio.\n')
    return dfclean

def rotinaseventos(outws, ws, asset_sn):
    site_name = findsitename(asset_sn)
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)
    df = replace_text(df, events_replace_text)
    df['Code'] = df['Code'].astype(str)
    
    df = delalerts(df)
        
    df['Asset'] = asset_sn
    df['Site'] = site_name
    df.to_csv(outws, encoding='utf-8-sig', index=False)

def rotinaseventossum(outws2, ws2, asset_sn):
    site_name = findsitename(asset_sn)
    data2 = ws2.values
    columns = next(data2)[0:]
    dfs = pd.DataFrame(data2, columns=columns)
    dfs = replace_text(dfs, events_replace_text)

    if all(pd.Series(std_eventsum_list).isin(dfs.columns)):
        count_series = dfs.groupby(std_eventsum_list).size()
        eventsalerts = count_series.to_frame(name='Count').reset_index()

    else:
        eventsalerts = pd.DataFrame(columns=std_eventsum_list)

    eventsalerts['Code'] = eventsalerts['Code'].astype(str)
    eventsalerts['Asset'] = asset_sn
    eventsalerts['Site'] = site_name
    eventsalerts = delalerts(eventsalerts)
        
    eventsalerts.to_csv(outws2, encoding='utf-8-sig', index=False)

#####################################################################################################
#                                          CONVERSÃO DE EVENTOS                                     #
#####################################################################################################

def eventsconvert(eventfile, ts_file):
    global troublefile
    global asset_sn
    global outws
    global outws2

    eventfile = os.path.join(eventfile)

    troublefile = os.path.join(ts_file)

    workfolder = os.path.join(destinationfolder, eventsdirectory)
    workfolder2 = os.path.join(destinationfolder, eventssumdirectory)

    try:
        #Abre xlsx dos troubleshoot dos alertas
        ts_df1 = openpyxl.load_workbook(troublefile)
        sheet = ts_df1.active
        data = sheet.values
        columns = next(data)[0:]
        ts_df1 = pd.DataFrame(data, columns=columns)
        #---
        ts_df1 = ts_df1[['Código', 'Causas possíveis', 'Recomendações', 'Peso']]
        ts_df1.columns = ['Code', 'Causas', 'Recomendações', 'Peso']
        ts_df1.dropna(how='all', inplace=True)
    except FileNotFoundError:
        exit()

    checkdestiny(workfolder)
    try:
        cdf = pd.read_csv(destinationfolder + '/events_output.csv', low_memory=False)
        cdf.to_csv(workfolder + '/events_output.csv', encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=std_event_list)
        cdf.to_csv(workfolder + '/events_output.csv', encoding='utf-8-sig', index=False)

    checkdestiny(workfolder2)
    try:
        cdf1 = pd.read_csv(workfolder2 + '/eventssum_output.csv', low_memory=False)
    except FileNotFoundError:
        cdf1 = pd.DataFrame(columns=std_eventsum_list)
        cdf1.to_csv(workfolder2 + '/eventssum_output.csv', encoding='utf-8-sig', index=False)

    if eventfile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(eventfile)
        print(' ')
        print('Iniciando tratamento de dados de Eventos e Alertas...')
        print(' ')
        for a_name in asset_list:
            if a_name != 'Totals':
                asset_sn = a_name[-8:]
                outws = workfolder + '/' + asset_sn + '.csv'
                outws2 = workfolder2 + '/' + asset_sn + '.csv'
                try:
                    ws = wb[str(a_name)]
                    print(a_name, 'Ok')
                    rotinaseventos(outws, ws, asset_sn)
                    rotinaseventossum(outws2, ws, asset_sn)
                except KeyError:
                    print(a_name, 'Vazio')
                    ws = pd.DataFrame(columns=std_event_list)
                    ws['Asset'] = asset_sn
                    ws['Site'] = findsitename(asset_sn)
                    ws['Timestamp'] = pd.to_datetime(ws['Timestamp'])
                    ws['Code'] = ws['Code'].astype(str)
                    ws.to_csv(outws, encoding='utf-8-sig', index=False)

                    ws2 = pd.DataFrame(columns=std_eventsum_list)
                    ws2['Asset'] = asset_sn
                    ws2['Site'] = findsitename(asset_sn)
                    ws2['Code'] = ws2['Code'].astype(str)
                    ws2.to_csv(outws2, encoding='utf-8-sig', index=False)

        # CONCATENA TUDO NA PLANILHA FINAL
    
    concatenarev(workfolder, 'events_')
    evdfoutput = concatenar_profile(workfolder2, std_eventsum_list, 'eventssum_')
    
          
    
    print(' ')
    print('Iniciando aquisição de texto de troubleshoot e cálculo do índice de confiabilidade...')
    print(' ')
    # MANUTENÇÃO COMEÇA CALCULANDO CONSUMO POR DIA, TOTAL DE CONSUMO, HORAS DE SERVIÇO POR DIA, TOTAL DE HORAS DE SERVIÇO

    ######  ÍNDICE DE CONFIABILIDADE

    try:    
        evdfoutput['Code'] = evdfoutput['Code'].str.replace(':', '-')
        evdfoutput.loc[evdfoutput['Type'] == 'Event', 'IsEv'] = 'E'
        evdfoutput['CodeB'] = evdfoutput['Code'].replace("-.+", "", regex=True)
        evdfoutput['CodeB'] = evdfoutput['IsEv'] + evdfoutput['CodeB']
        evdfoutput.loc[evdfoutput['Type'] == 'Event', 'Code'] = evdfoutput.loc[evdfoutput['Type'] == 'Event', 'CodeB']
    
        if keepfiles == 0:
            evdfoutput.drop(columns=['IsEv', 'CodeB'], inplace=True)
    
        ts_df1.replace("/\(.+", "", inplace=True, regex=True)
    
        evdfoutput.dropna(how='all', inplace=True)
        evdfoutput = pd.merge_ordered(evdfoutput, ts_df1, how='left')
        evdfoutput['PesoSev'] = evdfoutput['Severity']
    
        evdfoutput['PesoSev'] = evdfoutput['PesoSev'].str.replace('Low', '5.6234')
        evdfoutput['PesoSev'] = evdfoutput['PesoSev'].str.replace('Medium', '10')
        evdfoutput['PesoSev'] = evdfoutput['PesoSev'].str.replace('High', '17.7828')
    
        evdfoutput['PesoSev'] = pd.to_numeric(evdfoutput['PesoSev'])
        evdfoutput['Peso'] = pd.to_numeric(evdfoutput['Peso'])
    
        evdfoutput['PesoReal'] = np.power(evdfoutput['PesoSev'], 4) * np.power(evdfoutput['Peso'], 2)
        evdfoutput['IC'] = ((evdfoutput['Count'] * evdfoutput['PesoReal']) / 10) / Pdias
    
        # ---
        # Definição do status de confiabilidade
        sconf2 = 133
        sconf3 = 333
        evdfoutput["Status de confiabilidade"] = "Verde"
    
        evdfoutput.loc[evdfoutput['IC'] >= sconf2, 'Status de confiabilidade'] = 'Amarelo'
        evdfoutput.loc[evdfoutput['IC'] >= sconf3, 'Status de confiabilidade'] = 'Vermelho'
        
        # ---
        
        evdfoutput.append(pd.DataFrame({'Severity': ['Low', 'Medium', 'High']}))
        
        outfs = workfolder2 + '/eventssum_output.csv'
        if keepfiles == 0:
            evdfoutput.drop(columns=['PesoSev', 'Peso', 'PesoReal'], inplace=True)
    except ValueError:
        outfs = workfolder2 + '/eventssum_output.csv'
    
    evdfoutput.to_csv(outfs, encoding='utf-8-sig', index=False)
    try:
        txt = open(destinationfolder + "/Events_html.txt", "w+")
        txt.write(evhtml(evdfoutput))
        txt.close()
    except:
        pass
    
    #fix Sem alertas de alguma severidade
    evsum = workfolder2 + 'eventssum_output.csv'
    evsumdf = pd.read_csv(evsum, low_memory=False)
    evsev = evsumdf['Severity'].unique()
    if not 'Low' in evsev:
        evsumdf = evsumdf.append({'Severity':'Low','Count':0},ignore_index=True)
    if not 'Medium' in evsev:
        evsumdf = evsumdf.append({'Severity':'Medium','Count':0},ignore_index=True)
    if not 'High' in evsev:
        evsumdf = evsumdf.append({'Severity':'High','Count':0},ignore_index=True)

    evsumdf = evsumdf.append(pd.DataFrame(columns=['Causas', 'Recomendações', 'IC', 'Status de confiabilidade']))
    evsumdf.to_csv(evsum, encoding='utf-8-sig', index=False)
    #--------
    

def evhtml(df: pd.core.frame.DataFrame):
    """ 
        Função que gera código em formatação HTML dos alertas.
  
        Parameters: 
            df (DataFrame): Caminho do arquivo Excel com os alertas que a serem descritos
            
        Returns: 
            Txhtml (srt): Descrição de alertas em HTML
        """
    df = df[["Code", "Description", "Causas", "Recomendações"]].drop_duplicates(subset=["Code"])
    df.reset_index(inplace=True)
    Txhtml = ""
    i = -1

    for VNome in df["Description"]:
        i += 1
        t1 = VNome
        t2 = df.loc[i, 'Causas']
        t3 = df.loc[i, 'Recomendações']
        Txhtml = Txhtml + "<b>" + str(t1) + "</b><br>" + "<b>Causas possíveis: </b>" + str(
            t2) + "<br><b>Recomendações: </b>" + str(t3) + "<br><br>"
    return Txhtml.replace('\n', '')

#####################################################################################################
#                               CONVERSÃO DE HISTÓRICOS DE DADOS                                    #
#####################################################################################################

def historyconvert(historyfile):
    global asset_sn
    global dfoutput
    global fcd, fct, shd, sht
    global dfo, dfm
    dfm = pd.DataFrame()

    historyfile = os.path.join(historyfile)

    workfolder = os.path.join(destinationfolder, historydirectory)

    checkdestiny(workfolder)
    try:
        cdf = pd.read_csv(destinationfolder + '/history_output.csv', low_memory=False)
        cdf.to_csv(workfolder + '/history_output.csv', encoding='utf-8-sig', index=False)
        
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=std_param_list)
        cdf.to_csv(workfolder + '/history_output.csv', encoding='utf-8-sig', index=False)

    dailyfolder = os.path.join(destinationfolder, dailysumdir)
    checkdestiny(dailyfolder)

    try:
        cdf = pd.read_csv(destinationfolder + '/historyday_output.csv', low_memory=False)
        cdf.to_csv(dailyfolder + '/historyday_output.csv', encoding='utf-8-sig', index=False)
    except FileNotFoundError:
        cdf = pd.DataFrame(columns=std_param_list)
        cdf.to_csv(dailyfolder + '/historyday_output.csv', encoding='utf-8-sig', index=False)
    
    ############################################################################
    ##################   CASO O ARQUIVO DE ENTRADA SEJA ZIP   ##################
    ############################################################################
    global Pdias, ws
    
    if historyfile.endswith('.zip'):
        zf = zipfile.ZipFile(historyfile)
        print(' ')
        print('Iniciando Extracting, Transforming and Loading...')
        print(' ')
        
        
        if (not os.path.exists(os.path.dirname(historyfile) + '/BackupDataLog.zip')) and (os.path.basename(destinationfolder) == '01 - BD_CARLAO' or 
                                                                                          os.path.basename(destinationfolder) == '01 - BD_CARGILL'):
            shutil.copyfile(historyfile, os.path.dirname(historyfile) + '/BackupDataLog.zip')
        
        for snFull in asset_list:
            sn = snFull[-8:]
            dfcarlao = dfmodule.loc[dfmodule['SN']==sn]
            dfcarlao = dfcarlao.loc[dfmodule['Modulo']=='carlao']
            if len(dfcarlao) > 0:
                zf = carlao.carlao(snFull,historyfile)
                asset_list.append('MCA - D1K01363')
        
        for snFull in asset_list:
            sn = snFull[-8:]
            dfcargill = dfmodule.loc[dfmodule['SN']==sn]
            dfcargill = dfcargill.loc[dfmodule['Modulo']=='cargill']
            if len(dfcargill) > 0:
                zf = cargill.cargill(snFull,historyfile)
                if sn == 'S2K00384':    
                    asset_list.extend(['S1M06675', 'S1M06677'])
                elif sn == 'S2K00386':
                    asset_list.extend(['S1M06672', 'S1M06678'])

        global dictsmhday, dictfuelday
        dictsmhday = {}
        dictfuelday = {}
        
        for a_name in asset_list:
            if a_name != 'Totals':
                ws = a_name + '.csv'

                asset_sn = a_name[-8:]
                try:
                    dataframe = pd.read_csv(zf.open(ws), encoding='utf-16le', dtype=object)
                    
                    dataframe.dropna(how='all', axis=1, inplace=True)
                    
                    dataframe = renamecol(dataframe, asset_sn)
                    
                    if contacol(dataframe, r'.*Run Hours.*') == 0:
                        
                        if dataframe.columns.to_list().count('Total Operating Hours [Hours]') > 0:
                           dataframe.rename(columns={'Total Operating Hours [Hours]': 'Run Hours'}, inplace=True)
                    if contacol(dataframe, r'.*Run Hours.*') == 0:        
                        if dataframe.columns.to_list().count('Engine Total Hours of Operation [Hrs]') > 0:
                           dataframe.rename(columns={'Engine Total Hours of Operation [Hrs]': 'Run Hours'}, inplace=True)
                    if contacol(dataframe, r'.*Run Hours.*') == 0:     
                        if dataframe.columns.to_list().count('Total Operating Hours [Hrs]') > 0:
                           dataframe.rename(columns={'Total Operating Hours [Hrs]': 'Run Hours'}, inplace=True)
                           
                    
                    if contacol(dataframe, r'.*Run Hours.*') == 0:
                        
                        if dataframe.columns.to_list().count('Total Time [Hours]') > 0:
                           dataframe.rename(columns={'Total Time [Hours]': 'Run Hours'}, inplace=True)
                    if contacol(dataframe, r'.*Run Hours.*') == 0:       
                        if dataframe.columns.to_list().count('Total Time [Hrs]') > 0:
                           dataframe.rename(columns={'Total Time [Hrs]': 'Run Hours'}, inplace=True)
                    
                    if contacol(dataframe, r'.*Engine Load.*') == 0:
                        
                        if dataframe.columns.to_list().count('Engine Percent Load At Current Speed [%]') > 0:
                           dataframe.rename(columns={'Engine Percent Load At Current Speed [%]': 'Engine Load'}, inplace=True)
                           
                    if contacol(dataframe, r'.*Fuel Consumption Rate.*') == 0:
                        
                        if dataframe.columns.to_list().count('Engine Fuel Rate [L/hr]') > 0:
                           dataframe.rename(columns={'Engine Fuel Rate [L/hr]': 'Fuel Consumption Rate'}, inplace=True)
           
                    if contacol(dataframe, r'.*Boost Pressure.*') == 0:
                        
                        if dataframe.columns.to_list().count('Engine Intake Manifold #1 Pressure [kPa]') > 0:
                           dataframe.rename(columns={'Engine Intake Manifold #1 Pressure [kPa]': 'Boost Pressure'}, inplace=True)
                    
                    if contacol(dataframe, r'.*Engine Speed.*') == 0:
                        
                        if dataframe.columns.to_list().count('Engine\'s Desired Operating Speed [RPM]') > 0:
                           dataframe.rename(columns={'Engine\'s Desired Operating Speed [RPM]': 'Engine Speed'}, inplace=True)
                    
                    if contacol(dataframe, r'.*Left Exhaust Temp.*') == 0:
                        
                        if dataframe.columns.to_list().count('Engine Exhaust Manifold Bank 1 Temperature 1 [Deg. C]') > 0:
                           dataframe.rename(columns={'Engine Exhaust Manifold Bank 1 Temperature 1 [Deg. C]': 'Left Exhaust Temp'}, inplace=True)
                    
                    if contacol(dataframe, r'.*Right Exhaust Temp.*') == 0:
                        
                        if dataframe.columns.to_list().count('Engine Exhaust Manifold Bank 2 Temperature 1 [Deg. C]') > 0:
                           dataframe.rename(columns={'Engine Exhaust Manifold Bank 2 Temperature 1 [Deg. C]': 'Right Exhaust Temp'}, inplace=True)
                    
                    #print('agora ',asset_sn)
                    if contacol(dataframe, r'Total Fuel \[L\]') == 0:
                        
                        if dataframe.columns.to_list().count('Engine Total Fuel Used [L]') > 0:
                           dataframe.rename(columns={'Engine Total Fuel Used [L]': 'Total Fuel'}, inplace=True)
                           
                    if contacol(dataframe, r'.*Oil Pressure \[kPa\].*') == 0:
                        
                        if dataframe.columns.to_list().count('Engine Oil Pressure 1 [kPa]') > 0:
                           dataframe.rename(columns={'Engine Oil Pressure 1 [kPa]': 'Oil Pressure'}, inplace=True)
                           
                        
                    dataframe = delcol(dataframe)
                    
                    dataframe = replace_text(dataframe, data_replace_text)
                    
                  
                    dataframe['Timestamp'] = pd.to_datetime(dataframe['Timestamp'])

                except KeyError:
                    dataframe = pd.DataFrame(columns=std_param_list)
                    dataframe['Timestamp'] = pd.to_datetime(dataframe['Timestamp'])

                # chama rotinas aplicáveis ao dataframe
                rotinas(dataframe)
                # salva dataframe pronto na pasta dos parametros
                dfclean.to_csv(os.path.join(destinationfolder, workfolder) + asset_sn + '.csv', encoding='utf-8-sig',
                               index=False)

                dictsmhday[asset_sn] = smhcalc(dfclean, asset_sn)
                dictfuelday[asset_sn] = fuelcalc(dfclean, asset_sn)
                
                dfhday = dfclean
                dfhday.reset_index(drop=True, inplace=True)
                dfhday.set_index(pd.DatetimeIndex(dfhday['Timestamp']), inplace=True)
                dfhday = dfhday.resample('D').mean()
                dfhday = dfhday.replace(0, np.nan)
                dfhday['Asset'] = asset_sn
                dfhday['Site'] = findsitename(asset_sn)
                dailyfile = os.path.join(dailyfolder) + '/' + asset_sn + '.csv'
                dfhday.to_csv(dailyfile, encoding='utf-8-sig', index=True)

        # CONCATENA TUDO NA PLANILHA FINAL
        dfoutput = concatenar(os.path.join(destinationfolder, workfolder), 'history_')

        Pdias = np.timedelta64(
            pd.to_datetime(dfoutput['Timestamp']).max() - pd.to_datetime(dfoutput['Timestamp']).min(), 'D').astype(int)

        concatenar_profile(os.path.join(destinationfolder, loadhistdir), std_loadhist_list, 'loadhist_')
        concatenar_profile(os.path.join(destinationfolder, rpmhistdir), std_rpmhist_list, 'rpmhist_')
        concatenar(dailyfolder, 'historyday_')

        print(' ')
        print('Iniciando calculos de intervalos de manutenção...')
        print(' ')
        # MANUTENÇÃO COMEÇA CALCULANDO CONSUMO POR DIA, TOTAL DE CONSUMO, HORAS DE SERVIÇO POR DIA, TOTAL DE HORAS DE SERVIÇO
        try:
            dfo = pd.read_csv(os.path.join(destinationfolder) + '/maintenance_output.csv')
        except FileNotFoundError:
            dfo = pd.DataFrame(columns=maintenance_list)
        for a_name in asset_list:
            if a_name != 'Totals':
                asset_sn = a_name[-8:]
                dfa = dfoutput.query('Asset == @asset_sn')
                lastused = pd.to_datetime(dfa['Timestamp']).max()
                datasetvazio = 0
                if dfa.empty:
                    datasetvazio = 1
                # print(asset_sn + ' - Dataset Vazio: ' + str(datasetvazio))
                dfm = maintenanceoutput(dfoutput, lastused, asset_sn, datasetvazio)
                dfo = pd.concat([dfo, dfm])
                dfo.drop_duplicates(subset=['Asset'], inplace=True, keep='last')
                dfo.dropna(how='all', axis=0, inplace=True)
        checkdestiny(os.path.join(destinationfolder, mandirectory))
        dfo.to_csv(os.path.join(destinationfolder, mandirectory) + 'maintenance_output.csv', encoding='utf-8-sig',
                   index=False)

        ####### UTILIZAÇÃO #######
        print(' ')
        print('Iniciando estudos de Utilização e Taxas de Consumo de Combustível...')
        print(' ')

        outdir1 = os.path.join(destinationfolder, loadstudydir)
        outdir1h = os.path.join(destinationfolder, loadstudydirh)
        outdir2 = os.path.join(destinationfolder, loadresumedir)
        outdir3 = os.path.join(destinationfolder, fuelstudydir)
        outdir4 = os.path.join(destinationfolder, fuelresumedir)
        checkdestiny(outdir1)
        checkdestiny(outdir1h)
        checkdestiny(outdir2)
        checkdestiny(outdir3)
        checkdestiny(outdir4)

        for sit in sites_list:
            print('Sítio:', sit)
            slist = findsiteassets(sit)
            count = 0
            salist = []
            for a_name in slist:
                if count < 1:
                    df1 = dfoutput.query('Asset == @a_name')[['Timestamp', 'Load', 'Fuel_Rate']]
                    df1.columns = ['Timestamp' if x == 'Timestamp' else str(a_name) + '_' + x for x in df1.columns]
                    df1.reset_index(drop=True, inplace=True)
                    dfh = df1
                    salist.append(a_name)
                else:
                    df1 = dfoutput.query('Asset == @a_name')[['Timestamp', 'Load', 'Fuel_Rate']]
                    df1.columns = ['Timestamp' if x == 'Timestamp' else str(a_name) + '_' + x for x in df1.columns]
                    df1.reset_index(drop=True, inplace=True)
                    if 'Timestamp' in dfh.columns:
                        dfh = pd.merge_ordered(dfh, df1, fill_method='None')
                    salist.append(a_name)
                count += 1

            # dfh.dropna(how='all', axis=1, inplace=True)
            dfh.reset_index(drop=True, inplace=True)
            dfh.set_index(pd.DatetimeIndex(dfh['Timestamp']), inplace=True)

            dfhhr = dfh.resample('1H').mean()
            dfhhr = dfhhr.replace(0, np.nan)

            dfhmin = dfh.resample('5Min').mean()
            dfhmin = dfhmin.replace(0, np.nan)

            print('Lista de Ativos:', salist)

            # LOAD RESUME
            loadtable = pd.concat([dfhhr.loc[:, dfhhr.columns.isin(['Timestamp'])],
                                   dfhhr.loc[:, dfhhr.columns.str.contains('Load')]], axis=1)

            loadtable['QNT_SIM'] = loadtable.apply(lambda x: x.notnull().sum(), axis='columns')

            fin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(slist) + 1)], ignore_index=True)
            l = []
            for i in range(len(slist) + 1):
                l.append(loadtable[loadtable['QNT_SIM'] == (i)].count()['QNT_SIM'] / len(loadtable['QNT_SIM']) * 100)
            fin['%'] = l
            fin['Site'] = sit
            finfile = os.path.join(outdir2) + '/' + sit + '_LOAD_RESUME' + '.csv'
            fin.to_csv(finfile, encoding='utf-8-sig', index=False)

            # LOAD STUDY
            ltvd = loadtable.resample('1D').mean().round(0)
            for x in range(len(loadtable.columns) - 1):
                ltvd['UT' + str(x + 1) + 'E'] = (
                            (loadtable['QNT_SIM'] == (x + 1)).resample('1D').sum().astype(int) / 24 * 100).round(1)
            if keepfiles == 0:
                ltvd.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvd['Site'] = sit
            finfile2 = os.path.join(outdir1) + '/' + sit + '_LOAD_STUDY' + '.csv'
            ltvd.to_csv(finfile2, encoding='utf-8-sig', index=True)

            # LOAD STUDY 1h
            loadtablemin = pd.concat([dfhmin.loc[:, dfhmin.columns.isin(['Timestamp'])],
                                      dfhmin.loc[:, dfhmin.columns.str.contains('Load')]], axis=1)
            loadtablemin['QNT_SIM'] = loadtablemin.apply(lambda x: x.notnull().sum(), axis='columns')

            ltvdh = loadtablemin.resample('1H').mean().round(0)
            for x in range(len(loadtablemin.columns) - 1):
                ltvdh['UT' + str(x + 1) + 'E'] = (
                            (loadtablemin['QNT_SIM'] == (x + 1)).resample('1H').sum().astype(int) / 12 * 100).round(1)
            if keepfiles == 0:
                ltvdh.drop(['QNT_SIM'], axis=1, inplace=True)
            ltvdh['Site'] = sit
            finfile2h = os.path.join(outdir1h) + '/' + sit + '_LOAD_STUDY_H' + '.csv'
            ltvdh.to_csv(finfile2h, encoding='utf-8-sig', index=True)

            # FUEL RESUME
            fratetable = pd.concat([dfhhr.loc[:, dfhhr.columns.isin(['Timestamp'])],
                                    dfhhr.loc[:, dfhhr.columns.str.contains('Fuel_Rate')]], axis=1)

            rtvb = fratetable.copy()
            rtvb['AVG'] = rtvb.sum(numeric_only=True, axis=1)

            rtvfile = os.path.join(outdir3) + '/' + sit + '_FR_STUDY' + '.csv'
            fratetable = fratetable.replace(0, np.nan)
            rtvD = fratetable.resample('1D').mean().round(0)
            rtvD['Site'] = sit
            rtvD.to_csv(rtvfile, encoding='utf-8-sig', index=True)

            fratetable['QNT_SIM'] = fratetable.apply(lambda x: x.notnull().sum(), axis='columns')

            rin = pd.concat([pd.DataFrame([i], columns=['QNT']) for i in range(len(slist) + 1)], ignore_index=True)
            l = []
            for i in range(len(slist) + 1):
                l.append(rtvb[fratetable['QNT_SIM'] == (i)].mean()['AVG'])
            rin['L/hr'] = l
            rin['Site'] = sit
            rinfile = os.path.join(outdir4) + '/' + sit + '_FR_RESUME' + '.csv'
            rin.to_csv(rinfile, encoding='utf-8-sig', index=False)

        concatenar_study(outdir1, 'loadstudy_')
        concatenar_study(outdir1h, 'loadstudy_H_')
        concatenar_study(outdir2, 'loadresume_')
        concatenar_study(outdir3, 'fuelstudy_')
        concatenar_study(outdir4, 'fuelresume_')

#####################################################################################################
#                                           ENVIROMENT                                              #
#####################################################################################################


def preplistas(engine_file, event_file, ts_file, destfolder, deb, concatenardb):

    global scriptname
    scriptname = os.path.basename(__file__)

    global swdir
    # swdir = os.getcwd()
    swdir = (Path(destfolder).parent)
    print(swdir)

    global destinationfolder
    destinationfolder = destfolder

    global keepfiles
    keepfiles = deb

    global historydirectory
    historydirectory = 'History/'
    global eventsdirectory
    eventsdirectory = 'Events/'
    global eventssumdirectory
    eventssumdirectory = 'EventsSummary/'
    global mandirectory
    mandirectory = 'Maintenance/'
    global rpmhistdir
    rpmhistdir = 'RPMProfile/'
    global loadhistdir
    loadhistdir = 'LoadProfile/'
    global loadresumedir
    loadresumedir = 'LoadResume/'
    global fuelresumedir
    fuelresumedir = 'FuelResume/'
    global loadstudydir
    loadstudydir = 'LoadStudy/'
    global loadstudydirh
    loadstudydirh = 'LoadStudy_H/'
    global fuelstudydir
    fuelstudydir = 'FuelStudy/'
    global dailysumdir
    dailysumdir = 'DailySummary/'

    global infodir
    infodir = os.path.join(swdir, '00 - INFOS/')
    assetinfofile = infodir + 'ASSET_INFO.xlsx'
    global assetlistdf
    assetlistdf = openfilewb(assetinfofile,'ASSET_LIST')
    global dirconfig, dflistparmraw, dflistcoldel, dflistalertdel, list_invalid_data, dfmainte
    dirconfig = infodir + 'ConfigScript.xlsx'
    
    dflistinvdata = openfilewb(dirconfig,'DadosInvalidos')
    list_invalid_data = dflistinvdata.values.tolist()
    
    dflistparmraw =  openfilewb(dirconfig,'ListaParm')
    dflistcoldel = openfilewb(dirconfig,'ColunasDelete')
    dflistalertdel = openfilewb(dirconfig,'AlertasDelete')
    try:
        dfmainte = openfilewb(dirconfig,'Maintenance')
    except:
        dfmainte = pd.DataFrame(columns=['SN','Método'])

    global asset_list
    asset_list = getlistativos(event_file)
    asset_list = asset_list.loc[asset_list!='MEZ00103']
    asset_list = asset_list.to_list()
    global sites_list
    sites_list = getlistasites(assetinfofile)
    sites_list = [i for i in sites_list if i is not None]
    global dfmodule
    try:
        dfmodule = openfilewb(dirconfig,'Modulos')
    except KeyError:
        dfmodule = pd.DataFrame(columns=['SN','Modulo'])
        
        
    global f_density
    f_density = 850
    
    if concatenardb == 0:
        for filename in os.listdir(destinationfolder):
            file_path = os.path.join(destinationfolder, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.remove(file_path)
                elif os.path.isdir(file_path):
                    rmtree(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))

    destinationfolder = os.path.join(destinationfolder)
    checkdestiny(destinationfolder)
    historyconvert(engine_file)
    eventsconvert(event_file, ts_file)
    limpadao(destinationfolder)
    
    print('Finished RFV 2.0 Handling Script V5.4')
    print()



if __name__ == '__main__':
    print('Você deve executar o arquivo GUI.py do conversor.')
    exit()

